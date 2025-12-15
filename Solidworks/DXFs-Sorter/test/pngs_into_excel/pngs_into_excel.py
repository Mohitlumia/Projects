import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from image_utils import scale_image_proportionally, center_image_in_cell

# --- USER SETTINGS ---
excel_path = r"D:\Solidworks\Bunker\Bunker BOM\Bunker BOM r2 - Copy.xlsx"
images_folder = r"D:\Solidworks\Bunker\Bunker BOM\Bunker_DXFs_PNGs"
part_number_column = "J"  # Modify if your Part Number column is different
image_column = "L"        # Column where image will be inserted
image_size = (100, 100)   # Resize width, height (optional)

# --- SCRIPT STARTS ---
wb = load_workbook(excel_path)
ws = wb.active

for row in range(2, ws.max_row + 1):  # assume row 1 = header
    part_num = ws[f"{part_number_column}{row}"].value

    if not part_num:
        continue

    image_path = os.path.join(images_folder, f"{part_num}.png")

    if os.path.exists(image_path):
        img = Image(image_path)
        scale_image_proportionally(img, max_width=120, max_height=120)  # change bounding size as needed
        
        # Convert pixel → Excel column width & row height
        # Excel column width is approx. 7 pixels per width unit
        column_width = img.width / 7
        row_height = img.height * 0.75  # Excel row height uses ~0.75 px per unit

        col_letter = image_column
        row_number = row

        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, column_width)
        ws.row_dimensions[row_number].height = max(ws.row_dimensions[row_number].height, row_height)


        ws.add_image(img, f"{image_column}{row}") # Insert image into cell

        # center_image_in_cell(ws, img, image_column, row) # Center image in cell
    else:
        print(f"⚠️ Image not found for: {part_num}")

# Save Excel with images
output_path = excel_path.replace(".xlsx", "_with_images 4.xlsx")
wb.save(output_path)

print("🎯 Done! Images inserted successfully.")
print(f"📁 Output saved as: {output_path}")
